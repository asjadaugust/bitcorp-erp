"""Tests para servicio de exportación Excel."""

from io import BytesIO

from openpyxl import load_workbook

from app.servicios.exportar import ServicioExportar


def test_exportar_generar_excel() -> None:
    """Debe generar un archivo Excel con datos."""
    servicio = ServicioExportar()
    datos = [
        {"id": 1, "nombre": "Test 1", "estado": "ACTIVO"},
        {"id": 2, "nombre": "Test 2", "estado": "INACTIVO"},
    ]
    columnas = ["id", "nombre", "estado"]
    resultado = servicio.generar_excel(datos, columnas, "Prueba")

    assert isinstance(resultado, BytesIO)
    wb = load_workbook(resultado)
    ws = wb.active
    assert ws.title == "Prueba"
    assert ws.max_row == 3  # Header + 2 data rows
    assert ws.max_column == 3


def test_exportar_equipos() -> None:
    """Debe exportar equipos con columnas correctas."""
    servicio = ServicioExportar()
    datos = [
        {
            "id": 1,
            "codigo": "EQ-001",
            "denominacion": "Excavadora",
            "placa": "ABC-123",
            "estado": "ACTIVO",
            "tipo_equipo_nombre": "MAQUINARIA_PESADA",
            "proveedor_nombre": "Proveedor X",
            "created_at": "2026-01-01",
        }
    ]
    resultado = servicio.exportar_equipos(datos)
    wb = load_workbook(resultado)
    ws = wb.active
    assert ws.title == "Equipos"
    assert ws.max_row == 2  # Header + 1 data row


def test_exportar_valorizaciones() -> None:
    """Debe exportar valorizaciones con columnas correctas."""
    servicio = ServicioExportar()
    datos = [
        {
            "id": 1,
            "codigo": "VAL-001",
            "contrato_codigo": "CON-001",
            "periodo": "2026-03",
            "estado": "BORRADOR",
            "monto_total": 5000.00,
            "created_at": "2026-01-01",
        }
    ]
    resultado = servicio.exportar_valorizaciones(datos)
    wb = load_workbook(resultado)
    ws = wb.active
    assert ws.title == "Valorizaciones"
    assert ws.max_row == 2
