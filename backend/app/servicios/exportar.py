"""Servicio de exportación Excel.
"""

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.config.logging import obtener_logger

logger = obtener_logger(__name__)

# Header style
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="003F87", end_color="003F87", fill_type="solid")


class ServicioExportar:
    """Servicio para exportación de datos a Excel."""

    def generar_excel(
        self,
        datos: list[dict[str, Any]],
        columnas: list[str],
        nombre_hoja: str = "Datos",
    ) -> BytesIO:
        """Generar un archivo Excel genérico a partir de datos."""
        wb = Workbook()
        ws = wb.active
        ws.title = nombre_hoja  # type: ignore[union-attr]

        # Headers
        for col_idx, col_name in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)  # type: ignore[union-attr]
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL

        # Data rows
        for row_idx, fila in enumerate(datos, 2):
            for col_idx, col_name in enumerate(columnas, 1):
                ws.cell(row=row_idx, column=col_idx, value=fila.get(col_name, ""))  # type: ignore[union-attr]

        # Auto-width columns
        for col_idx, col_name in enumerate(columnas, 1):
            max_len = len(str(col_name))
            for fila in datos:
                val = str(fila.get(col_name, ""))
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = ws.cell(row=1, column=col_idx).column_letter  # type: ignore[union-attr]
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        logger.info("excel_generado", filas=len(datos), columnas=len(columnas))
        return output

    def exportar_equipos(self, datos: list[dict[str, Any]]) -> BytesIO:
        """Exportar lista de equipos a Excel."""
        columnas = [
            "id", "codigo", "denominacion", "placa", "estado",
            "tipo_equipo_nombre", "proveedor_nombre", "created_at",
        ]
        return self.generar_excel(datos, columnas, "Equipos")

    def exportar_valorizaciones(self, datos: list[dict[str, Any]]) -> BytesIO:
        """Exportar lista de valorizaciones a Excel."""
        columnas = [
            "id", "codigo", "contrato_codigo", "periodo", "estado",
            "monto_total", "created_at",
        ]
        return self.generar_excel(datos, columnas, "Valorizaciones")
